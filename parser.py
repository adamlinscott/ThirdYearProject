import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from string import ascii_uppercase
import os

wb = load_workbook('xl.xlsx')
sheetNames = wb.get_sheet_names()
jsonFile = open('input.json','w')
jsonFile.write('{\n')
ws = wb.active

#find total weeks in tornament
totalWeeks = 0
for i in range(1, 50):
	cell = 'A'+str(i)
	if type(ws[cell].value) == type(long()):
		if int(ws[cell].value) > totalWeeks:
			totalWeeks = ws[cell].value
jsonFile.write('\t"total_weeks":' + str(totalWeeks) + ',\n')


#find starting date
startDate = 0
for i in range(1, totalWeeks):
	cell = 'A'+str(i)
	if ws[cell].value == 1:
		startDate = ws['C'+str(i)].value
		break
jsonFile.write('\t"start_date":"' + str(startDate) + '",\n')


#find holiday dates
jsonFile.write('\t"bank_hols":[\n')
holCellStart = 0
for i in range(totalWeeks, totalWeeks+20):
	cell = 'A'+str(i)
	if isinstance(ws[cell].value, unicode):
		holCellStart = i
		break
if holCellStart == 0:
	jsonFile.write('\t],\n')
else:
	jsonFile.write('\t\t"' + str(ws['C'+str(holCellStart)].value) + '"')
	for i in range(holCellStart, holCellStart+20):
		if isinstance(ws['C'+str(i+1)].value, datetime):
			jsonFile.write(',\n\t\t"' + str(ws['C'+str(i+1)].value) + '"')
		else:
			jsonFile.write('\n\t],\n')
			break

			

jsonFile.write('\t"divisions":[\n')
#loop to read data per division
for name in sheetNames:
	ws = wb[name] 
	name = name.replace(" ", "_")
	#open division object
	jsonFile.write('\t{\n')
	jsonFile.write('\t\t"name":"' + name + '",\n')
	
	#find first week for division
	for i in range(1, totalWeeks):
		if type(ws['B'+str(i)].value) == type(long()):
			jsonFile.write('\t\t"start_week":' + str(ws[ 'A'+str(i)].value) + ',\n')
			break
	
	#find team information
	teamInfoZeroIndex = 0
	for c in range(ord('D'),ord('Z')):
		if isinstance(ws[str(unichr(c)) + '1'].value, unicode):
			teamInfoZeroIndex = str(unichr(c-1))
			break
	numOfTeams = 0
	for i in range(1, 40):
		if type(ws[teamInfoZeroIndex + str(i)].value) == type(long()):
			if int(ws[teamInfoZeroIndex + str(i)].value) > numOfTeams:
				numOfTeams = ws[teamInfoZeroIndex + str(i)].value
	jsonFile.write('\t\t"team_num":' + str(numOfTeams) + ',\n')
	
	currentTeam = 1
	jsonFile.write('\t\t"teams":[\n')
	
	jsonFile.write('\t\t\t{\n')
	jsonFile.write('\t\t\t\t"num":1,\n')
	jsonFile.write('\t\t\t\t"name":"' + ws[str(unichr(ord(teamInfoZeroIndex)+1)) + "2"].value + '",\n') 
	
	jsonFile.write('\t\t\t\t"requirements":[') 
	isRequirements = False
	for j in range(2, totalWeeks+2):
#		if isinstance(ws[unichr(ord('C') + 1) + str(j)].value, unicode):
			if isinstance(ws['A' + str(j)].value, long):
		#		isRequirements = True
				jsonFile.write('\n\t\t\t\t\t"' + str(ws[unichr(ord('C') + 1) + str(j)].value) + '",')
#	jsonFile.write('\n\t\t\t\t\t"bank_hols":{\n')
#	jsonFile.write('\t\t\t\t\t}\n')
	jsonFile.write('\n\t\t\t\t\t"None"')
	jsonFile.write('\n\t\t\t\t],\n')
	jsonFile.write('\t\t\t\t"grounds":' + "null" + '\n') #Must change when implementing ground sharing conditions
				
	jsonFile.write('\t\t\t}')
	
	for i in range(2, numOfTeams+1):
		jsonFile.write(',\n\t\t\t{\n')
		jsonFile.write('\t\t\t\t"num":' + str(i) + ',\n')
		jsonFile.write('\t\t\t\t"name":"' + ws[str(unichr(ord(teamInfoZeroIndex)+1)) + str(i+1)].value + '",\n') 
		jsonFile.write('\t\t\t\t"requirements":[') 
		isRequirements = False
		for j in range(2, totalWeeks+2):
#			if isinstance(ws[unichr(ord('C') + i) + str(j)].value, unicode):
				if isinstance(ws['A' + str(j)].value, long):
		#			isRequirements = True
					jsonFile.write('\n\t\t\t\t\t"' + str(ws[unichr(ord('C') + i) + str(j)].value) + '",')
#		jsonFile.write('\n\t\t\t\t\t"bank_hols":{\n')
#		jsonFile.write('\t\t\t\t\t}\n')
		jsonFile.write('\n\t\t\t\t\t"None"')
		jsonFile.write('\n\t\t\t\t],\n')
		jsonFile.write('\t\t\t\t"grounds":' + "null" + '\n') #Must change when implementing ground sharing conditions
		jsonFile.write('\t\t\t}')
	jsonFile.write('\n\t\t]\n')
	#close division object
	jsonFile.write('\t},\n')
#end of name in sheetNames for loop

jsonFile.write('\t],\n')


jsonFile.write('\t"file_parsed":true\n')
jsonFile.write('}\n')
#close file
jsonFile.close()